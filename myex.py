import  openpyxl as xl

# exf는 엑셀파일을 불러오는 객체/변수/핸들러
# 임의로 이름을 만들고, 파이썬<->엑셀의 매개체 역할담당
exf = xl.load_workbook('c:\\dd\\ess.xlsx')

aws = exf.active
#엑셀의 자료를 불러오기
for i in aws.rows:
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    mat = i[3].value
#불러온 자료를 수정하기
    tot = kor + eng + mat
    avg = tot / 3
    
    aws.cell(row = index, column = 5).value = tot
    aws.cell(row = index, column = 6).value = avg


    print('{} {} {} {} {} {:.2f} '.format(name, kor, eng, mat, tot, avg))
#수정된 값을 엑셀로 출력
exf.save('c:\\dd\\outss.xlsx')
exf.close()



#1 openpyxl as xl = 엑셀 로드 함수 선언 as xl
#2 xl 객체로 엑셀 파일 불러오는 객체 exf 선언
#3 exf를 통해 불러온 파일을 활성화시켜 aws라는 객체로 선언

#4 aws(활성화된 엑셀파일)의 row(행)의 자료를 i번 만큼 반복하여 불러옴
#5 0번부터 3번까지 불러올 value의 위치를 지정함

#6 자료를 원하는대로 수정하고 aws의 column에 위치 지정하여 데이터 입력
#7 이름을 바꿔서 save
